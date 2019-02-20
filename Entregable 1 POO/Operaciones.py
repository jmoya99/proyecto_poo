import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from string import ascii_uppercase
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import Alignment
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from arrow import utcnow
from Mensajes import Mensajes as Me

"""
    Correo: saan.unal@gmail.com
    Contraseña: SAAN12345
"""
# Template de excel tomado de:
# Nombre:       reporteEXCEL.py
# Autor:        Miguel Andres Garcia Niño
# Creado:       03 de Agosto 2018
# Modificado:   03 de Agosto 2018
# Copyright:    (c) 2018 by Miguel Andres Garcia Niño, 2018
# License:      Apache License 2.0


class Operaciones:

    @staticmethod
    def exportar_excel(titulo, cabecera, registros, nombre_excel):
        # Workbook es el contenedor para todas las demás partes del documento.
        libro_trabajo = Workbook()

        hoja = libro_trabajo.active
        hoja.title = (titulo)
        hoja.sheet_properties.tabColor = "1072BA"

        # Ver líneas de cuadrícula
        hoja.sheet_view.showGridLines = False

        celda_final = ascii_uppercase[len(cabecera)]
        rango_titulo = "B2:{}3".format(celda_final)
        rango_cabecera = "B10:{}10".format(celda_final)

        centrar_texto = Alignment(horizontal="center", vertical="center")

      # ========================== TÍTULO ==========================

        hoja.merge_cells(rango_titulo)
        celda_titulo = hoja.cell(row=2, column=2)
        celda_titulo.value = titulo.upper()
        celda_titulo.alignment = centrar_texto
        celda_titulo.font = Font(color="FF000000", size=11, bold=True)

      # ===================== INFORMACIÓN EXTRA ====================

        fuente_informacion_extra = Font(color="707070", size=11, bold=False)

        celda_origen = hoja.cell(row=5, column=2)
        celda_origen.value = "Generado por SAAN"
        celda_origen.font = fuente_informacion_extra

        celda_fecha_descarga = hoja.cell(row=6, column=2)
        celda_fecha_descarga.value = "Fecha de descarga: {}".format(
            utcnow().to("local").format("DD/MM/YYYY"))
        celda_fecha_descarga.font = fuente_informacion_extra

        celda_cantidad_descarga = hoja.cell(row=8, column=2)
        celda_cantidad_descarga.value = "Registros descargados: {}".format(
            len(registros))
        celda_cantidad_descarga.font = fuente_informacion_extra

      # ================== BORDES - COLOR (CELDAS) =================

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        color_celda = PatternFill("solid", fgColor="C0C0C0")

      # ================== BORDES - COLOR (TÍTULO) =================

        filas_titulo = hoja[rango_titulo]

        celda_inicial = filas_titulo[0][0].row
        for fila in filas_titulo:
            fila_izquierda = fila[0]
            fila_derecha = fila[-1]
            fila_izquierda.border = fila_izquierda.border + \
                Border(left=border.left)
            fila_derecha.border = fila_derecha.border + \
                Border(right=border.right)

            for celda in fila:
                if celda.row == celda_inicial:
                    celda.border = celda.border + Border(top=border.top)
                else:
                    celda.border = celda.border + Border(bottom=border.bottom)

                celda.fill = color_celda

      # ========= DATOS - BORDES - COLOR (CABECERA - TABLA) ========

        for indice, dato in enumerate(cabecera, start=2):
            hoja.cell(row=10, column=indice).value = dato
            hoja.cell(row=10, column=indice).border = border
            hoja.cell(row=10, column=indice).alignment = centrar_texto
            hoja.cell(row=10, column=indice).font = Font(
                color="FF000000", size=10, bold=True)

        filas_encabezado = hoja[rango_cabecera]
        for fila in filas_encabezado:
            for celda in fila:
                celda.fill = color_celda

      # ====== REGISTROS - BORDES - COLOR (REGISTROS - TABLA) ======

        for fila_indice, registros in enumerate(registros, start=11):
            for columna_indice, registro in enumerate(registros, start=2):
                hoja.cell(row=fila_indice,
                          column=columna_indice).value = registro
                hoja.cell(row=fila_indice,
                          column=columna_indice).border = border
                hoja.cell(row=fila_indice, column=columna_indice).alignment = Alignment(horizontal="left",
                                                                                        vertical="center")
                hoja.cell(row=fila_indice, column=columna_indice).font = Font(color="FF000000",
                                                                              size=10, bold=False)

      # ============== AJUSTAR ANCHO (CELDAS - TABLA) ==============

        for col in hoja.columns:
            columna = [(columna.column, columna.value) for columna in col
                       if not columna.value is None]
            if columna:
                longitud_maxima = 0
                for celda in columna:
                    if len(str(celda[1])) > longitud_maxima:
                        longitud_maxima = len(celda[1])

                ajustar_ancho = (longitud_maxima + 1) * 1.2
                hoja.column_dimensions[columna[0][0]].width = ajustar_ancho

        try:
            # Guardar el libro actual bajo el nombre de archivo dado
            libro_trabajo.save("{}.xlsx".format(nombre_excel))

         # +----------------------------------------+
            retornar = Me.mensa["reporte"]
         # +----------------------------------------+
        except PermissionError:
         # +------------------------------------------------+
            retornar = Me.mensa["reportErr"]
         # +------------------------------------------------+
        except:
         # +-------------------------------+
            retornar = Me.mensa["err"]
         # +-------------------------------+
        finally:
            # Cerrar el libro de trabajo (Workbook)
            libro_trabajo.close()

            return retornar

    @staticmethod
    def enviar_correo_electronico(correo_enviar, asunto, cuerpo):
        server = smtplib.SMTP(host='smtp-relay.gmail.com', port=587)
        msg = MIMEMultipart()
        password = "SAAN12345"
        msg['From'] = "saan.unal@gmail.com"
        msg['To'] = correo_enviar
        msg['Subject'] = asunto
        msg.attach(MIMEText(cuerpo, 'plain'))
        server = smtplib.SMTP('smtp.gmail.com:587')
        server.starttls()
        server.login(msg['From'], password)
        server.sendmail(msg['From'], msg['To'], msg.as_string())
        server.quit()
